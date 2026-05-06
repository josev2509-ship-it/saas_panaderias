import json
from io import BytesIO
from datetime import datetime, date, timedelta
from decimal import Decimal
from urllib.parse import urlencode

from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .decorators import modulo_requerido

from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.utils import ImageReader

from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings

try:
    import qrcode
except ImportError:
    qrcode = None

from .models import (
    Empresa,
    CentroEducativo,
    MenuDiario,
    Conduce,
    ProductoFacturacion,
    ComprobanteFiscal,
    RangoComprobanteGubernamental,
    Factura,
    DetalleFactura,
    Plan,
    EmpresaSaaS,
    Suscripcion,
    PerfilUsuario,
    CodigoValidacion,
)

from .utils import (
    generar_pdf_conduce,
    generar_pdf_conduces_masivo,
    generar_pdf_relacion_diaria,
)


# =====================================================
# FUNCIONES AUXILIARES
# =====================================================

def convertir_fecha(fecha_str):
    if not fecha_str:
        return None

    try:
        return datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        return None


def convertir_fecha_excel(fecha):
    if isinstance(fecha, datetime):
        return fecha.date()

    if isinstance(fecha, date):
        return fecha

    if isinstance(fecha, (int, float)):
        return from_excel(fecha).date()

    if isinstance(fecha, str):
        fecha_limpia = fecha.strip()

        for formato in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(fecha_limpia, formato).date()
            except ValueError:
                pass

    return None


def obtener_empresa(request):
    """
    Devuelve la empresa operativa asociada al usuario autenticado.
    Esto evita que un usuario vea datos de otra empresa.
    """
    if not request.user.is_authenticated:
        return None

    empresa = getattr(request.user, "empresa_principal", None)

    if empresa:
        return empresa

    perfil = PerfilUsuario.objects.filter(user=request.user).first()

    if perfil and perfil.empresa:
        empresa, creada = Empresa.objects.get_or_create(
            usuario=request.user,
            defaults={
                "nombre": perfil.empresa.nombre,
                "rnc": perfil.empresa.rnc,
                "correo": perfil.empresa.correo,
                "numero_inicial_conduce": "0001",
            }
        )
        return empresa

    return None


def fecha_corta(fecha):
    return f"{fecha.day}/{fecha.month}/{fecha.year}"


def fecha_larga_es(fecha):
    meses = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }

    return f"{fecha.day:02d} de {meses[fecha.month]} de {fecha.year}"


def formato_monto(valor):
    if valor is None:
        valor = Decimal("0.00")

    return f"{Decimal(valor):,.2f}"


def formato_cantidad(valor):
    if valor is None:
        return "0"

    return str(int(valor))


def formatear_fecha_grafico(valor):
    if not valor:
        return ""

    if isinstance(valor, datetime):
        valor = valor.date()

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    return str(valor)


def formatear_mes_grafico(valor):
    meses = {
        1: "Ene",
        2: "Feb",
        3: "Mar",
        4: "Abr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dic",
    }

    if not valor:
        return ""

    if isinstance(valor, datetime):
        valor = valor.date()

    if isinstance(valor, date):
        return f"{meses.get(valor.month, '')} {valor.year}"

    return str(valor)


def nombre_mes(fecha_inicio, fecha_fin):
    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    if fecha_inicio.month == fecha_fin.month and fecha_inicio.year == fecha_fin.year:
        return f"{meses[fecha_inicio.month]} {fecha_inicio.year}"

    return f"{fecha_inicio.strftime('%d/%m/%Y')} AL {fecha_fin.strftime('%d/%m/%Y')}"


def normalizar_producto(nombre):
    if not nombre:
        return "Otros"

    nombre = nombre.lower()

    if "muffin" in nombre or "mufin" in nombre:
        return "Bizcocho"

    if "bizcocho" in nombre or "biscocho" in nombre:
        return "Bizcocho"

    if "pan de zanahoria" in nombre:
        return "Pan con vegetales"

    if "vegetales" in nombre:
        return "Pan con vegetales"

    if "galleta" in nombre:
        return "Galleta"

    if "pan" in nombre:
        return "Pan"

    return "Otros"


def clasificar_producto(producto, cantidad):
    producto = (producto or "").upper().strip()

    pan = ""
    pan_vegetales = ""
    galleta = ""
    bizcocho = ""

    if "PAN DE ZANAHORIA" in producto:
        pan_vegetales = cantidad
    elif "MUFFIN" in producto or "MUFIN" in producto:
        bizcocho = cantidad
    elif "GALLETA" in producto or "GALLETAS" in producto:
        galleta = cantidad
    elif "BIZCOCHO" in producto or "BISCOCHO" in producto:
        bizcocho = cantidad
    elif "VEGETALES" in producto:
        pan_vegetales = cantidad
    else:
        pan = cantidad

    return pan, pan_vegetales, galleta, bizcocho


def clasificar_categoria_factura(producto):
    producto = (producto or "").upper().strip()

    if "PAN DE ZANAHORIA" in producto:
        return "PAN_CON_VEGETALES"

    if "MUFFIN" in producto or "MUFIN" in producto:
        return "BIZCOCHO"

    if "BIZCOCHO" in producto or "BISCOCHO" in producto:
        return "BIZCOCHO"

    if "GALLETA" in producto or "GALLETAS" in producto:
        return "GALLETA"

    if "VEGETALES" in producto:
        return "PAN_CON_VEGETALES"

    return "PAN"


# =====================================================
# DASHBOARD / PANEL PRINCIPAL
# =====================================================

@login_required(login_url="login_usuario")
def inicio(request):
    empresa = obtener_empresa(request)

    if not empresa:
        messages.error(request, "Debe configurar su empresa antes de continuar.")
        return redirect("login_usuario")

    hoy = timezone.localdate()
    manana = hoy + timedelta(days=1)
    inicio_mes = hoy.replace(day=1)

    conduces_mes = Conduce.objects.filter(
        empresa=empresa,
        fecha__gte=inicio_mes,
        fecha__lte=hoy
    )

    todos_conduces = Conduce.objects.filter(empresa=empresa)

    total_raciones = conduces_mes.aggregate(total=Sum("cantidad"))["total"] or 0
    total_conduces = conduces_mes.count()
    total_centros = CentroEducativo.objects.filter(empresa=empresa).count()

    menu_manana = MenuDiario.objects.filter(empresa=empresa, fecha=manana).first()
    producto_manana = menu_manana.producto if menu_manana else "No hay menú registrado"
    fecha_manana = manana.strftime("%d/%m/%Y")

    precio_racion_estimado = Decimal("10.18")
    proyeccion_ventas = f"{Decimal(total_raciones) * precio_racion_estimado:,.2f}"

    raciones_por_dia = (
        conduces_mes
        .values("fecha")
        .annotate(total=Sum("cantidad"))
        .order_by("fecha")
    )

    labels_dias = [formatear_fecha_grafico(item["fecha"]) for item in raciones_por_dia if item["fecha"]]
    data_dias = [item["total"] or 0 for item in raciones_por_dia if item["fecha"]]

    productos = {}

    for conduce in conduces_mes:
        producto = normalizar_producto(conduce.producto)
        productos[producto] = productos.get(producto, 0) + (conduce.cantidad or 0)

    labels_productos = list(productos.keys())
    data_productos = list(productos.values())

    resumen_meses = (
        todos_conduces
        .annotate(mes=TruncMonth("fecha"))
        .values("mes")
        .annotate(
            total_raciones=Sum("cantidad"),
            total_conduces=Count("id")
        )
        .order_by("-mes")
    )

    return render(request, "inicio.html", {
        "empresa": empresa,
        "total_raciones": total_raciones,
        "total_conduces": total_conduces,
        "total_centros": total_centros,
        "producto_manana": producto_manana,
        "fecha_manana": fecha_manana,
        "proyeccion_ventas": proyeccion_ventas,
        "labels_dias": labels_dias,
        "data_dias": data_dias,
        "labels_productos": labels_productos,
        "data_productos": data_productos,
        "resumen_meses": resumen_meses,
    })


# =====================================================
# PLANTILLAS EXCEL
# =====================================================

@login_required(login_url="login_usuario")
def descargar_plantilla_centros(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Centros"

    ws.append([
        "codigo",
        "nombre",
        "director",
        "telefono",
        "direccion",
        "provincia",
        "regional_distrito",
        "matricula",
        "latitud",
        "longitud",
    ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="plantilla_centros.xlsx"'
    wb.save(response)
    return response


@login_required(login_url="login_usuario")
def descargar_plantilla_menu(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    ws.append(["fecha", "producto"])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="plantilla_menu.xlsx"'
    wb.save(response)
    return response


# =====================================================
# GESTIÓN DE CENTROS
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def pantalla_carga_centros(request):
    empresa = obtener_empresa(request)
    q = request.GET.get("q", "").strip()

    centros = CentroEducativo.objects.filter(empresa=empresa).order_by("orden_carga", "id")

    if q:
        centros = centros.filter(
            Q(codigo__icontains=q) |
            Q(nombre__icontains=q) |
            Q(director__icontains=q) |
            Q(provincia__icontains=q) |
            Q(regional_distrito__icontains=q)
        )

    total_centros = centros.count()
    centros_con_ubicacion = centros.exclude(latitud__isnull=True).exclude(longitud__isnull=True).count()

    return render(request, "carga_centros.html", {
        "empresa": empresa,
        "centros": centros,
        "q": q,
        "total_centros": total_centros,
        "centros_con_ubicacion": centros_con_ubicacion,
    })


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def crear_centro(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        codigo = request.POST.get("codigo", "").strip()
        nombre = request.POST.get("nombre", "").strip()

        if not codigo or not nombre:
            messages.error(request, "El código y el nombre del centro son obligatorios.")
            return redirect("carga_centros")

        existe = CentroEducativo.objects.filter(empresa=empresa, codigo=codigo).exists()
        if existe:
            messages.error(request, "Ya existe un centro con ese código para esta empresa.")
            return redirect("carga_centros")

        CentroEducativo.objects.create(
            empresa=empresa,
            codigo=codigo,
            nombre=nombre,
            director=request.POST.get("director", "").strip(),
            telefono=request.POST.get("telefono", "").strip(),
            direccion=request.POST.get("direccion", "").strip(),
            provincia=request.POST.get("provincia", "").strip(),
            regional_distrito=request.POST.get("regional_distrito", "").strip(),
            matricula=int(request.POST.get("matricula") or 0),
            latitud=request.POST.get("latitud", "").strip().replace(",", ".") or None,
            longitud=request.POST.get("longitud", "").strip().replace(",", ".") or None,
            orden_carga=CentroEducativo.objects.filter(empresa=empresa).count() + 1,
        )

        messages.success(request, "Centro creado correctamente.")

    return redirect("carga_centros")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def editar_centro(request, centro_id):
    empresa = obtener_empresa(request)
    centro = get_object_or_404(CentroEducativo, id=centro_id, empresa=empresa)

    if request.method == "POST":
        centro.codigo = request.POST.get("codigo", "").strip()
        centro.nombre = request.POST.get("nombre", "").strip()
        centro.director = request.POST.get("director", "").strip()
        centro.telefono = request.POST.get("telefono", "").strip()
        centro.direccion = request.POST.get("direccion", "").strip()
        centro.provincia = request.POST.get("provincia", "").strip()
        centro.regional_distrito = request.POST.get("regional_distrito", "").strip()
        centro.matricula = int(request.POST.get("matricula") or 0)
        centro.latitud = request.POST.get("latitud", "").strip().replace(",", ".") or None
        centro.longitud = request.POST.get("longitud", "").strip().replace(",", ".") or None
        centro.save()

        messages.success(request, "Centro actualizado correctamente.")
        return redirect("carga_centros")

    return render(request, "editar_centro.html", {"empresa": empresa, "centro": centro})


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def eliminar_centro(request, centro_id):
    empresa = obtener_empresa(request)
    centro = get_object_or_404(CentroEducativo, id=centro_id, empresa=empresa)

    if request.method == "POST":
        centro.delete()
        messages.success(request, "Centro eliminado correctamente.")

    return redirect("carga_centros")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def cargar_centros_excel(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        archivo = request.FILES.get("archivo")

        if not archivo:
            messages.error(request, "Debe seleccionar un archivo.")
            return redirect("carga_centros")

        wb = load_workbook(archivo)
        ws = wb.active

        for indice, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            datos = list(row)

            while len(datos) < 10:
                datos.append(None)

            codigo, nombre, director, telefono, direccion, provincia, regional_distrito, matricula, latitud, longitud = datos[:10]

            if codigo and nombre:
                CentroEducativo.objects.update_or_create(
                    empresa=empresa,
                    codigo=str(codigo).strip(),
                    defaults={
                        "nombre": str(nombre).strip(),
                        "director": str(director or ""),
                        "telefono": str(telefono or ""),
                        "direccion": str(direccion or ""),
                        "provincia": str(provincia or ""),
                        "regional_distrito": str(regional_distrito or ""),
                        "matricula": int(matricula or 0),
                        "latitud": str(latitud).replace(",", ".") if latitud else None,
                        "longitud": str(longitud).replace(",", ".") if longitud else None,
                        "orden_carga": indice,
                    },
                )

        messages.success(request, "Centros cargados correctamente.")
        return redirect("carga_centros")

    return redirect("carga_centros")


# =====================================================
# MAPA DE CENTROS
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def mapa_centros(request):
    empresa = obtener_empresa(request)
    codigo = request.GET.get("codigo", "").strip()

    centros_con_ubicacion = (
        CentroEducativo.objects
        .filter(empresa=empresa)
        .exclude(latitud__isnull=True)
        .exclude(longitud__isnull=True)
        .order_by("codigo")
    )

    centro_buscado = None

    if codigo:
        centro_buscado = CentroEducativo.objects.filter(empresa=empresa, codigo__iexact=codigo).first()

    centros_json = []

    for centro in centros_con_ubicacion:
        centros_json.append({
            "codigo": str(centro.codigo),
            "nombre": str(centro.nombre),
            "director": centro.director or "",
            "telefono": centro.telefono or "",
            "direccion": centro.direccion or "",
            "provincia": centro.provincia or "",
            "distrito": centro.regional_distrito or "",
            "matricula": centro.matricula or 0,
            "latitud": float(centro.latitud),
            "longitud": float(centro.longitud),
        })

    mensaje_ubicacion = None

    if centro_buscado:
        if centro_buscado.latitud and centro_buscado.longitud:
            mensaje_ubicacion = "Este centro ya tiene ubicación registrada y se muestra en el mapa."
        else:
            mensaje_ubicacion = "Este centro no tiene ubicación registrada. Puedes agregar latitud y longitud."

    return render(request, "mapa_centros.html", {
        "empresa": empresa,
        "centros": centros_con_ubicacion,
        "centro_buscado": centro_buscado,
        "codigo": codigo,
        "centros_json": json.dumps(centros_json),
        "mensaje_ubicacion": mensaje_ubicacion,
    })


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_centros")
def actualizar_ubicacion_centro(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        centro_id = request.POST.get("centro_id")
        latitud = request.POST.get("latitud", "").strip().replace(",", ".")
        longitud = request.POST.get("longitud", "").strip().replace(",", ".")

        centro = get_object_or_404(CentroEducativo, id=centro_id, empresa=empresa)

        if not latitud or not longitud:
            messages.error(request, "Debe completar latitud y longitud.")
            return redirect(f"/centros/mapa/?codigo={centro.codigo}")

        centro.latitud = latitud
        centro.longitud = longitud
        centro.save()

        messages.success(request, f"Ubicación agregada correctamente al centro {centro.codigo}.")
        return redirect(f"/centros/mapa/?codigo={centro.codigo}")

    return redirect("mapa_centros")


# =====================================================
# GESTIÓN DE MENÚ
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_menu")
def pantalla_carga_menu(request):
    empresa = obtener_empresa(request)
    q = request.GET.get("q", "").strip()

    menus = MenuDiario.objects.filter(empresa=empresa).order_by("-fecha")

    if q:
        menus = menus.filter(
            Q(producto__icontains=q) |
            Q(fecha__icontains=q)
        )

    return render(request, "carga_menu.html", {
        "empresa": empresa,
        "menus": menus,
        "q": q,
    })


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_menu")
def crear_menu_diario(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        fecha = request.POST.get("fecha")
        producto = request.POST.get("producto", "").strip()

        if fecha and producto:
            MenuDiario.objects.update_or_create(
                empresa=empresa,
                fecha=fecha,
                defaults={"producto": producto}
            )
            messages.success(request, "Menú creado correctamente.")
        else:
            messages.error(request, "Debe completar la fecha y el producto.")

    return redirect("carga_menu")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_menu")
def editar_menu_diario(request, menu_id):
    empresa = obtener_empresa(request)
    menu = get_object_or_404(MenuDiario, id=menu_id, empresa=empresa)

    if request.method == "POST":
        fecha = request.POST.get("fecha")
        producto = request.POST.get("producto", "").strip()

        if fecha and producto:
            menu.fecha = fecha
            menu.producto = producto
            menu.save()
            messages.success(request, "Menú actualizado correctamente.")
            return redirect("carga_menu")

        messages.error(request, "Debe completar la fecha y el producto.")

    return render(request, "editar_menu.html", {"empresa": empresa, "menu": menu})


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_menu")
def eliminar_menu_diario(request, menu_id):
    empresa = obtener_empresa(request)
    menu = get_object_or_404(MenuDiario, id=menu_id, empresa=empresa)

    if request.method == "POST":
        menu.delete()
        messages.success(request, "Menú eliminado correctamente.")

    return redirect("carga_menu")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_menu")
def cargar_menu_excel(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        archivo = request.FILES.get("archivo")

        if not archivo:
            messages.error(request, "Debe seleccionar un archivo.")
            return redirect("carga_menu")

        wb = load_workbook(archivo)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha, producto = row
            fecha = convertir_fecha_excel(fecha)

            if fecha and producto:
                MenuDiario.objects.update_or_create(
                    empresa=empresa,
                    fecha=fecha,
                    defaults={"producto": str(producto).strip()},
                )

        messages.success(request, "Menú cargado correctamente.")
        return redirect("carga_menu")

    return redirect("carga_menu")


# =====================================================
# GENERACIÓN DE CONDUCES
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def generar_conduces_automaticos(request):
    empresa = obtener_empresa(request)
    empresas = Empresa.objects.filter(id=empresa.id)

    if request.method == "POST":
        numero_inicial = request.POST.get("numero_inicial", "").strip()
        fecha_desde = request.POST.get("fecha_desde")
        fecha_hasta = request.POST.get("fecha_hasta")

        if not fecha_desde or not fecha_hasta:
            messages.error(request, "Debe completar las fechas.")
            return redirect("generar_conduces")

        fecha_desde = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        fecha_hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()

        if fecha_desde > fecha_hasta:
            messages.error(request, "La fecha desde no puede ser mayor que la fecha hasta.")
            return redirect("generar_conduces")

        centros = CentroEducativo.objects.filter(empresa=empresa).order_by("orden_carga", "id")
        menus = MenuDiario.objects.filter(
            empresa=empresa,
            fecha__range=[fecha_desde, fecha_hasta]
        ).order_by("fecha")

        if not centros.exists():
            messages.error(request, "No hay centros cargados.")
            return redirect("generar_conduces")

        if not menus.exists():
            messages.error(request, "No hay menú cargado para ese rango de fechas.")
            return redirect("generar_conduces")

        def obtener_largo_formato(valor):
            return len(str(valor)) if str(valor).startswith("0") else None

        ultimo = Conduce.objects.filter(empresa=empresa).order_by("-id").first()

        if numero_inicial:
            numero = int(numero_inicial)
            largo = obtener_largo_formato(numero_inicial)
            empresa.numero_inicial_conduce = numero_inicial
            empresa.save()
        elif ultimo:
            numero = int(ultimo.numero) + 1
            largo = obtener_largo_formato(ultimo.numero)
        else:
            numero = int(empresa.numero_inicial_conduce or "1")
            largo = obtener_largo_formato(empresa.numero_inicial_conduce or "1")

        total_generados = 0

        for menu in menus:
            for centro in centros:
                existe = Conduce.objects.filter(
                    empresa=empresa,
                    fecha=menu.fecha,
                    centro=centro,
                ).exists()

                if not existe:
                    numero_final = str(numero).zfill(largo) if largo else str(numero)

                    Conduce.objects.create(
                        empresa=empresa,
                        numero=numero_final,
                        fecha=menu.fecha,
                        centro=centro,
                        producto=menu.producto,
                        cantidad=centro.matricula,
                        estado="borrador",
                    )

                    numero += 1
                    total_generados += 1

        messages.success(request, f"Se generaron {total_generados} conduces.")
        return redirect("generar_conduces")

    return render(request, "generar_conduces.html", {"empresa": empresa, "empresas": empresas})


# =====================================================
# BÚSQUEDA / EDICIÓN / ELIMINACIÓN DE CONDUCES
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def buscar_conduces(request):
    empresa = obtener_empresa(request)

    conduces = (
        Conduce.objects
        .select_related("centro", "empresa")
        .filter(empresa=empresa)
        .order_by("-fecha", "-id")
    )

    q = request.GET.get("q", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    estado = request.GET.get("estado", "").strip()

    if q:
        conduces = conduces.filter(
            Q(numero__icontains=q) |
            Q(producto__icontains=q) |
            Q(centro__nombre__icontains=q) |
            Q(centro__codigo__icontains=q)
        )

    if fecha_desde:
        conduces = conduces.filter(fecha__gte=fecha_desde)

    if fecha_hasta:
        conduces = conduces.filter(fecha__lte=fecha_hasta)

    if estado:
        conduces = conduces.filter(estado=estado)

    return render(request, "buscar_conduces.html", {
        "empresa": empresa,
        "conduces": conduces,
        "q": q,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "estado": estado,
    })


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def editar_conduce(request, conduce_id):
    empresa = obtener_empresa(request)
    conduce = get_object_or_404(Conduce, id=conduce_id, empresa=empresa)

    if request.method == "POST":
        conduce.numero = request.POST.get("numero")
        conduce.fecha = request.POST.get("fecha")
        conduce.producto = request.POST.get("producto")
        conduce.cantidad = request.POST.get("cantidad")
        conduce.estado = request.POST.get("estado")
        conduce.observaciones = request.POST.get("observaciones")
        conduce.save()

        messages.success(request, "Conduce actualizado correctamente.")
        return redirect("buscar_conduces")

    return render(request, "editar_conduce.html", {"empresa": empresa, "conduce": conduce})


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def eliminar_conduce(request, conduce_id):
    empresa = obtener_empresa(request)
    conduce = get_object_or_404(Conduce, id=conduce_id, empresa=empresa)

    if request.method == "POST":
        conduce.delete()
        messages.success(request, "Conduce eliminado correctamente.")

    return redirect("buscar_conduces")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def anular_conduce(request, conduce_id):
    empresa = obtener_empresa(request)
    conduce = get_object_or_404(Conduce, id=conduce_id, empresa=empresa)
    conduce.estado = "anulado"
    conduce.save()

    messages.success(request, "Conduce anulado correctamente.")
    return redirect("buscar_conduces")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def vista_conduce(request, conduce_id):
    empresa = obtener_empresa(request)
    conduce = get_object_or_404(Conduce, id=conduce_id, empresa=empresa)
    return render(request, "vista_conduce.html", {"empresa": empresa, "conduce": conduce})


# =====================================================
# ACCIONES MASIVAS
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def acciones_conduces(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        ids = request.POST.getlist("conduces")
        accion = request.POST.get("accion")

        if not ids:
            messages.error(request, "Debe seleccionar al menos un conduce.")
            return redirect("buscar_conduces")

        conduces_qs = (
            Conduce.objects
            .filter(empresa=empresa, id__in=ids)
            .select_related("centro", "empresa")
        )

        conduces = list(conduces_qs)

        conduces.sort(
            key=lambda c: (
                c.fecha,
                int(c.numero) if str(c.numero).isdigit() else 0,
                c.centro.orden_carga,
                c.centro.nombre,
            )
        )

        if accion == "anular":
            conduces_qs.update(estado="anulado")
            messages.success(request, "Conduces anulados correctamente.")
            return redirect("buscar_conduces")

        if accion == "entregado":
            conduces_qs.update(estado="entregado")
            messages.success(request, "Conduces marcados como entregados.")
            return redirect("buscar_conduces")

        if accion == "pdf_ver":
            pdf = generar_pdf_conduces_masivo(conduces)
            return FileResponse(pdf, content_type="application/pdf")

        if accion == "pdf_descargar":
            pdf = generar_pdf_conduces_masivo(conduces)
            response = FileResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="conduces_seleccionados.pdf"'
            return response

        if accion == "relacion_diaria_pdf":
            conduces_validos = [conduce for conduce in conduces if conduce.estado != "anulado"]
            pdf = generar_pdf_relacion_diaria(conduces_validos)
            return FileResponse(pdf, content_type="application/pdf")

    return redirect("buscar_conduces")


# =====================================================
# PDF CONDUCE INDIVIDUAL
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_conduces")
def visualizar_pdf_conduce(request, conduce_id):
    empresa = obtener_empresa(request)
    conduce = get_object_or_404(Conduce, id=conduce_id, empresa=empresa)
    archivo_pdf = generar_pdf_conduce(conduce)

    return FileResponse(open(archivo_pdf, "rb"), content_type="application/pdf")


# =====================================================
# RELACIÓN DIARIA
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_reportes")
def generar_relacion_diaria_pdf(request):
    empresa = obtener_empresa(request)
    fecha = request.GET.get("fecha")

    if not fecha:
        return HttpResponse("Debe enviar una fecha. Ejemplo: /relacion-diaria/pdf/?fecha=2026-03-02", status=400)

    fecha = convertir_fecha(fecha)

    if not fecha:
        return HttpResponse("Formato de fecha inválido. Use el formato YYYY-MM-DD.", status=400)

    conduces = (
        Conduce.objects
        .filter(empresa=empresa, fecha=fecha)
        .exclude(estado="anulado")
        .select_related("empresa", "centro")
        .order_by("numero")
    )

    if not conduces.exists():
        return HttpResponse("No hay conduces válidos para esa fecha.", status=404)

    pdf = generar_pdf_relacion_diaria(conduces)

    return FileResponse(pdf, content_type="application/pdf", filename="relacion_diaria.pdf")


# =====================================================
# RELACIÓN GENERAL
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_reportes")
def generar_relacion_general_pdf(request):
    empresa = obtener_empresa(request)

    if not empresa:
        return HttpResponse("Debe registrar una empresa antes de generar la relación general.", status=400)

    fecha_inicio = convertir_fecha(request.GET.get("fecha_inicio"))
    fecha_fin = convertir_fecha(request.GET.get("fecha_fin"))

    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Debe seleccionar fecha inicio y fecha final.", status=400)

    if fecha_inicio > fecha_fin:
        return HttpResponse("La fecha de inicio no puede ser mayor que la fecha final.", status=400)

    conduces = (
        Conduce.objects
        .filter(empresa=empresa, fecha__range=[fecha_inicio, fecha_fin])
        .exclude(estado="anulado")
        .select_related("centro", "empresa")
        .order_by("fecha", "numero")
    )

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    page_width, page_height = letter

    table_width = 540
    x_tabla = (page_width - table_width) / 2
    tabla_top_y = 600

    col_widths = [43, 50, 58, 210, 45, 55, 40, 39]
    filas_por_pagina = 38
    row_height = 13
    header_height = 30

    meses_texto = nombre_mes(fecha_inicio, fecha_fin)

    estilo_nombre = ParagraphStyle(name="NombreCentro", fontName="Helvetica", fontSize=5.4, leading=6, alignment=TA_LEFT)
    estilo_header = ParagraphStyle(name="Header", fontName="Helvetica-Bold", fontSize=5.2, leading=5.8, alignment=TA_CENTER)

    encabezados = [
        Paragraph("FECHA", estilo_header),
        Paragraph("NO. DE<br/>CONDUCE", estilo_header),
        Paragraph("CODIGO DEL<br/>CENTRO<br/>EDUCATIVO", estilo_header),
        Paragraph("NOMBRE DEL CENTRO EDUCATIVO", estilo_header),
        Paragraph("PAN", estilo_header),
        Paragraph("PAN CON<br/>VEGETALE", estilo_header),
        Paragraph("GALLETA", estilo_header),
        Paragraph("BIZCOCHO", estilo_header),
    ]

    filas = []
    total_pan = 0
    total_pan_vegetales = 0
    total_galleta = 0
    total_bizcocho = 0

    for conduce in conduces:
        cantidad = conduce.cantidad or 0
        pan, pan_vegetales, galleta, bizcocho = clasificar_producto(conduce.producto, cantidad)

        total_pan += pan if pan != "" else 0
        total_pan_vegetales += pan_vegetales if pan_vegetales != "" else 0
        total_galleta += galleta if galleta != "" else 0
        total_bizcocho += bizcocho if bizcocho != "" else 0

        filas.append([
            fecha_corta(conduce.fecha),
            str(conduce.numero),
            conduce.centro.codigo,
            Paragraph((conduce.centro.nombre or "").upper(), estilo_nombre),
            f"{pan:,}" if pan != "" else "",
            f"{pan_vegetales:,}" if pan_vegetales != "" else "",
            f"{galleta:,}" if galleta != "" else "",
            f"{bizcocho:,}" if bizcocho != "" else "",
        ])

    paginas_tabla = max(1, (len(filas) + filas_por_pagina - 1) // filas_por_pagina)
    total_paginas = paginas_tabla

    def dibujar_encabezado(pagina_actual):
        pdf.setFont("Helvetica-Bold", 8.8)
        pdf.drawCentredString(page_width / 2, 742, (empresa.nombre or "").upper())

        pdf.setFont("Helvetica", 7.2)
        pdf.drawCentredString(page_width / 2, 729, empresa.direccion or "")
        pdf.drawCentredString(page_width / 2, 718, f"Ciudad {empresa.ciudad or ''}")

        if empresa.correo:
            pdf.drawCentredString(page_width / 2, 707, f"Correo.: {empresa.correo}")
            pdf.drawCentredString(page_width / 2, 696, f"Telefono.: {empresa.telefono or ''}")
            pdf.drawCentredString(page_width / 2, 685, f"RNC.:{empresa.rnc or ''}")
            titulo_y = 660
            mes_y = 640
        else:
            pdf.drawCentredString(page_width / 2, 707, f"Telefono.: {empresa.telefono or ''}")
            pdf.drawCentredString(page_width / 2, 696, f"RNC.:{empresa.rnc or ''}")
            titulo_y = 665
            mes_y = 645

        pdf.setFont("Helvetica-Bold", 10.5)
        pdf.drawCentredString(page_width / 2, titulo_y, "RELACION DE CONDUCE")

        texto_mes = f"MES  {meses_texto}"
        pdf.setFont("Helvetica-Bold", 7.4)
        pdf.drawCentredString(page_width / 2, mes_y, texto_mes)

        ancho_mes = pdf.stringWidth(texto_mes, "Helvetica-Bold", 7.4)
        pdf.setLineWidth(0.5)
        pdf.line((page_width / 2) - (ancho_mes / 2), mes_y - 3, (page_width / 2) + (ancho_mes / 2), mes_y - 3)

        pdf.setFont("Helvetica", 7)
        pdf.drawCentredString(page_width / 2, 55, f"{pagina_actual} DE {total_paginas}")

    pagina = 1
    indice = 0
    ultima_y_tabla = tabla_top_y

    while indice < len(filas):
        dibujar_encabezado(pagina)

        bloque = filas[indice:indice + filas_por_pagina]
        data = [encabezados] + bloque
        row_heights = [header_height] + [row_height] * len(bloque)

        tabla = Table(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)

        tabla.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 5.2),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 5.4),
            ("ALIGN", (0, 1), (2, -1), "CENTER"),
            ("ALIGN", (4, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("LEADING", (0, 1), (-1, -1), 13),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))

        ancho_tabla, alto_tabla = tabla.wrap(0, 0)
        y_tabla = tabla_top_y - alto_tabla

        tabla.drawOn(pdf, x_tabla, y_tabla)
        ultima_y_tabla = y_tabla

        indice += filas_por_pagina

        if indice < len(filas):
            pdf.showPage()
            pagina += 1

    if not filas:
        dibujar_encabezado(pagina)

    y_total = ultima_y_tabla - 14

    if y_total < 90:
        y_total = 90

    total_data = [[
        "",
        "",
        "",
        "TOTAL",
        "-" if int(total_pan or 0) == 0 else f"{total_pan:,}",
        "-" if int(total_pan_vegetales or 0) == 0 else f"{total_pan_vegetales:,}",
        "-" if int(total_galleta or 0) == 0 else f"{total_galleta:,}",
        "-" if int(total_bizcocho or 0) == 0 else f"{total_bizcocho:,}",
    ]]

    tabla_total = Table(total_data, colWidths=col_widths, rowHeights=[14])

    tabla_total.setStyle(TableStyle([
        ("SPAN", (0, 0), (3, 0)),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 1.4, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 6.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    tabla_total.wrap(0, 0)
    tabla_total.drawOn(pdf, x_tabla, y_total)

    y_firma = y_total - 38

    if y_firma < 70:
        y_firma = 70

    pdf.setFont("Helvetica-Bold", 6.5)
    pdf.line(x_tabla + 20, y_firma, x_tabla + 195, y_firma)
    pdf.drawString(x_tabla + 30, y_firma - 12, "FIRMA Y SELLO DEL SUPLIDOR")

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="relacion_general_conduces.pdf"'
    return response


# =====================================================
# GESTIÓN DE FACTURACIÓN
# =====================================================

@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def facturacion(request):
    empresa = obtener_empresa(request)

    empresas = Empresa.objects.filter(id=empresa.id)
    productos = ProductoFacturacion.objects.filter(empresa=empresa).order_by("id")
    comprobantes = ComprobanteFiscal.objects.filter(empresa=empresa).order_by("ncf")
    comprobantes_disponibles = ComprobanteFiscal.objects.filter(empresa=empresa, usado=False).order_by("ncf")
    facturas = Factura.objects.select_related("empresa", "comprobante").filter(empresa=empresa).order_by("-fecha_factura", "-id")

    total_facturado = facturas.exclude(estado="anulada").aggregate(total=Sum("total"))["total"] or 0
    total_itbis = facturas.exclude(estado="anulada").aggregate(total=Sum("itbis"))["total"] or 0

    return render(request, "facturacion.html", {
        "empresa": empresa,
        "empresas": empresas,
        "productos": productos,
        "comprobantes": comprobantes,
        "comprobantes_disponibles": comprobantes_disponibles,
        "facturas": facturas,
        "total_facturado": total_facturado,
        "total_itbis": total_itbis,
    })


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def crear_producto_facturacion(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        categoria = request.POST.get("categoria")
        nombre_factura = request.POST.get("nombre", "").strip()
        precio = request.POST.get("precio_sin_itbis", "0").replace(",", ".")
        aplica_itbis = request.POST.get("aplica_itbis") == "on"
        porcentaje_itbis = request.POST.get("porcentaje_itbis", "18").replace(",", ".")

        ProductoFacturacion.objects.update_or_create(
            empresa=empresa,
            categoria=categoria,
            defaults={
                "nombre_factura": nombre_factura,
                "precio_sin_itbis": Decimal(precio),
                "aplica_itbis": aplica_itbis,
                "porcentaje_itbis": Decimal(porcentaje_itbis),
                "activo": True,
            }
        )

        messages.success(request, "Producto de facturación guardado correctamente.")

    return redirect("facturacion")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def editar_producto_facturacion(request, producto_id):
    empresa = obtener_empresa(request)
    producto = get_object_or_404(ProductoFacturacion, id=producto_id, empresa=empresa)

    if request.method == "POST":
        producto.categoria = request.POST.get("categoria")
        producto.nombre_factura = request.POST.get("nombre_factura", "").strip()
        producto.precio_sin_itbis = Decimal(request.POST.get("precio_sin_itbis", "0").replace(",", "."))
        producto.aplica_itbis = request.POST.get("aplica_itbis") == "on"
        producto.porcentaje_itbis = Decimal(request.POST.get("porcentaje_itbis", "18").replace(",", "."))
        producto.activo = request.POST.get("activo") == "on"
        producto.save()

        messages.success(request, "Producto actualizado correctamente.")
        return redirect("facturacion")

    return render(request, "editar_producto_facturacion.html", {"empresa": empresa, "producto": producto})


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def eliminar_producto_facturacion(request, producto_id):
    empresa = obtener_empresa(request)
    producto = get_object_or_404(ProductoFacturacion, id=producto_id, empresa=empresa)

    if request.method == "POST":
        producto.delete()
        messages.success(request, "Producto eliminado correctamente.")

    return redirect("facturacion")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def crear_comprobante_fiscal(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        tipo = request.POST.get("tipo", "B15").strip().upper()
        ncf = request.POST.get("ncf", "").strip().upper()
        fecha_validez = request.POST.get("fecha_validez")

        if not ncf or not fecha_validez:
            messages.error(request, "Debe completar el NCF/e-NCF y la fecha de validez.")
            return redirect("facturacion")

        if tipo != "OTRO" and not ncf.startswith(tipo):
            messages.error(request, f"El NCF/e-NCF debe iniciar con {tipo}. Ejemplo: {tipo}00000001.")
            return redirect("facturacion")

        comprobante, creado = ComprobanteFiscal.objects.get_or_create(
            empresa=empresa,
            ncf=ncf,
            defaults={
                "tipo": tipo,
                "fecha_validez": fecha_validez,
                "usado": False,
            }
        )

        if not creado:
            comprobante.tipo = tipo
            comprobante.fecha_validez = fecha_validez
            comprobante.save()
            messages.success(request, "Comprobante actualizado correctamente.")
        else:
            messages.success(request, "Comprobante registrado correctamente.")

    return redirect("facturacion")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def crear_rango_ncf(request):
    empresa = obtener_empresa(request)

    if request.method == "POST":
        tipo = request.POST.get("tipo", "B15").strip().upper()
        prefijo = request.POST.get("prefijo", tipo).strip().upper()
        desde = int(request.POST.get("desde"))
        hasta = int(request.POST.get("hasta"))
        fecha_validez = request.POST.get("fecha_validez")

        if desde > hasta:
            messages.error(request, "El rango indicado no es válido.")
            return redirect("facturacion")

        if tipo != "OTRO" and prefijo != tipo:
            messages.error(request, f"El prefijo debe coincidir con el tipo seleccionado: {tipo}.")
            return redirect("facturacion")

        RangoComprobanteGubernamental.objects.create(
            prefijo=prefijo,
            numero_desde=desde,
            numero_hasta=hasta,
            fecha_validez=fecha_validez,
        )

        creados = 0

        for numero in range(desde, hasta + 1):
            ncf = f"{prefijo}{str(numero).zfill(8)}"

            _, created = ComprobanteFiscal.objects.get_or_create(
                empresa=empresa,
                ncf=ncf,
                defaults={
                    "tipo": tipo,
                    "fecha_validez": fecha_validez,
                    "usado": False,
                }
            )

            if created:
                creados += 1

        messages.success(request, f"{creados} comprobantes creados correctamente.")

    return redirect("facturacion")


@transaction.atomic
@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def generar_factura(request):
    empresa = obtener_empresa(request)

    if request.method != "POST":
        return redirect("facturacion")

    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_fin")
    comprobante_id = request.POST.get("comprobante")
    bloques = int(request.POST.get("bloques") or 1)

    if not fecha_inicio or not fecha_fin:
        messages.error(request, "Debe completar fecha inicio y fecha fin.")
        return redirect("facturacion")

    fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

    conduces = (
        Conduce.objects
        .filter(empresa=empresa, fecha__range=[fecha_inicio, fecha_fin])
        .exclude(estado="anulado")
        .select_related("centro", "empresa")
        .order_by("fecha", "numero")
    )

    if not conduces.exists():
        messages.error(request, "No existen conduces válidos para ese período.")
        return redirect("facturacion")

    if comprobante_id:
        comprobante = get_object_or_404(ComprobanteFiscal, id=comprobante_id, empresa=empresa, usado=False)
    else:
        comprobante = ComprobanteFiscal.objects.filter(empresa=empresa, usado=False).order_by("ncf").first()

    if not comprobante:
        messages.error(request, "No hay comprobantes disponibles.")
        return redirect("facturacion")

    productos_config = {p.categoria: p for p in ProductoFacturacion.objects.filter(empresa=empresa, activo=True)}
    categorias_requeridas = ["PAN", "PAN_CON_VEGETALES", "GALLETA", "BIZCOCHO"]

    for categoria in categorias_requeridas:
        if categoria not in productos_config:
            messages.error(request, f"Falta configurar el producto de facturación: {categoria}.")
            return redirect("facturacion")

    cantidades = {"PAN": 0, "PAN_CON_VEGETALES": 0, "GALLETA": 0, "BIZCOCHO": 0}

    for conduce in conduces:
        categoria = clasificar_categoria_factura(conduce.producto)
        cantidades[categoria] += conduce.cantidad or 0

    conduces_lista = list(conduces)

    def numero_orden(conduce):
        try:
            return int(conduce.numero)
        except Exception:
            return 0

    conduces_lista.sort(key=numero_orden)

    conduce_inicial = conduces_lista[0].numero
    conduce_final = conduces_lista[-1].numero

    fecha_factura = max(c.fecha for c in conduces_lista)
    primera_entrega = min(c.fecha for c in conduces_lista)
    ultima_entrega = max(c.fecha for c in conduces_lista)

    factura = Factura.objects.create(
        empresa=empresa,
        comprobante=comprobante,
        fecha_factura=fecha_factura,
        fecha_inicio=primera_entrega,
        fecha_fin=ultima_entrega,
        cantidad_conduces=len(conduces_lista),
        conduce_inicial=conduce_inicial,
        conduce_final=conduce_final,
        bloques=bloques,
        estado="emitida",
    )

    subtotal_exento = Decimal("0.00")
    subtotal_gravado = Decimal("0.00")

    for categoria, cantidad in cantidades.items():
        producto_config = productos_config[categoria]
        precio = producto_config.precio_sin_itbis
        valor = Decimal(cantidad) * precio

        DetalleFactura.objects.create(
            factura=factura,
            producto=producto_config.nombre_factura,
            categoria=categoria,
            cantidad=cantidad,
            precio_sin_itbis=precio,
            aplica_itbis=producto_config.aplica_itbis,
            valor=valor,
        )

        if producto_config.aplica_itbis:
            subtotal_gravado += valor
        else:
            subtotal_exento += valor

    subtotal = subtotal_exento + subtotal_gravado

    porcentaje_itbis = Decimal("18.00")
    productos_gravados = ProductoFacturacion.objects.filter(empresa=empresa, activo=True, aplica_itbis=True)

    if productos_gravados.exists():
        porcentaje_itbis = productos_gravados.first().porcentaje_itbis

    itbis = subtotal_gravado * (porcentaje_itbis / Decimal("100"))
    total = subtotal + itbis

    factura.subtotal_exento = subtotal_exento
    factura.subtotal_gravado = subtotal_gravado
    factura.subtotal = subtotal
    factura.itbis = itbis
    factura.total = total
    factura.save()

    comprobante.usado = True
    comprobante.fecha_uso = timezone.localdate()
    comprobante.save()

    messages.success(request, f"Factura generada correctamente con NCF {comprobante.ncf}.")

    return redirect("facturacion")


@transaction.atomic
@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def editar_factura(request, factura_id):
    empresa = obtener_empresa(request)

    factura = get_object_or_404(
        Factura.objects.select_related("comprobante", "empresa"),
        id=factura_id,
        empresa=empresa
    )

    comprobantes_disponibles = ComprobanteFiscal.objects.filter(empresa=empresa, usado=False).order_by("ncf")

    if request.method == "POST":
        factura.bloques = int(request.POST.get("bloques", 1))
        factura.estado = request.POST.get("estado", factura.estado)

        comprobante_id = request.POST.get("comprobante")

        if comprobante_id:
            nuevo = get_object_or_404(ComprobanteFiscal, id=comprobante_id, empresa=empresa)

            if factura.comprobante and factura.comprobante.id != nuevo.id:
                factura.comprobante.usado = False
                factura.comprobante.fecha_uso = None
                factura.comprobante.save()

            factura.comprobante = nuevo
            nuevo.usado = True
            nuevo.fecha_uso = timezone.localdate()
            nuevo.save()

        factura.es_electronica = request.POST.get("es_electronica") == "on"
        factura.encf = request.POST.get("encf") or None
        factura.codigo_seguridad = request.POST.get("codigo_seguridad") or None
        factura.url_qr = request.POST.get("url_qr") or None

        fecha_firma = request.POST.get("fecha_firma_digital")

        if fecha_firma:
            factura.fecha_firma_digital = datetime.strptime(fecha_firma, "%Y-%m-%dT%H:%M")
        else:
            factura.fecha_firma_digital = None

        factura.save()

        messages.success(request, "Factura actualizada correctamente.")
        return redirect("facturacion")

    return render(request, "editar_factura.html", {
        "empresa": empresa,
        "factura": factura,
        "comprobantes_disponibles": comprobantes_disponibles,
    })


@transaction.atomic
@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def anular_factura(request, factura_id):
    empresa = obtener_empresa(request)

    factura = get_object_or_404(
        Factura.objects.select_related("comprobante"),
        id=factura_id,
        empresa=empresa
    )

    if request.method == "POST":
        factura.estado = "anulada"
        factura.save()

        messages.success(request, "Factura anulada correctamente. El NCF permanece utilizado.")

    return redirect("facturacion")


@transaction.atomic
@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def eliminar_factura(request, factura_id):
    empresa = obtener_empresa(request)

    factura = get_object_or_404(
        Factura.objects.select_related("comprobante"),
        id=factura_id,
        empresa=empresa
    )

    if request.method == "POST":
        comprobante = factura.comprobante

        if comprobante:
            comprobante.usado = False
            comprobante.fecha_uso = None
            comprobante.save()

        factura.delete()

        messages.success(request, "Factura eliminada. El NCF quedó disponible nuevamente.")

    return redirect("facturacion")


def construir_url_qr_ecf(factura):
    empresa = factura.empresa
    comprobante = factura.comprobante

    encf = factura.encf or (comprobante.ncf if comprobante else "")

    if not factura.codigo_seguridad or not factura.fecha_firma_digital:
        return ""

    parametros = {
        "RncEmisor": empresa.rnc or "",
        "RncComprador": factura.cliente_rnc or "",
        "ENCF": encf,
        "FechaEmision": factura.fecha_factura.strftime("%d-%m-%Y"),
        "MontoTotal": formato_monto(factura.total).replace(",", ""),
        "FechaFirma": factura.fecha_firma_digital.strftime("%d-%m-%Y %H:%M:%S"),
        "CodigoSeguridad": factura.codigo_seguridad,
    }

    return "https://ecf.dgii.gov.do/ecf/ConsultaTimbre?" + urlencode(parametros)


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def pdf_factura(request, factura_id):
    empresa_usuario = obtener_empresa(request)

    factura = get_object_or_404(
        Factura.objects.select_related("empresa", "comprobante"),
        id=factura_id,
        empresa=empresa_usuario
    )

    detalles = factura.detalles.all().order_by("id")

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    empresa = factura.empresa
    comprobante = factura.comprobante

    margin_left = 58
    margin_right = 58
    content_width = width - margin_left - margin_right

    x_left = margin_left
    x_right = 382

    y = 710

    pdf.setFont("Helvetica", 8)
    pdf.drawString(x_left, y, (empresa.nombre or "").upper())

    y -= 12
    pdf.drawString(x_left, y, empresa.direccion or "")

    y -= 12
    pdf.drawString(x_left, y, f"Ciudad {empresa.ciudad or ''}")

    y -= 12
    pdf.drawString(x_left, y, f"Teléfono  {empresa.telefono or ''}")

    y -= 12
    pdf.drawString(x_left, y, f"RNC-{empresa.rnc or ''}")

    y -= 12
    pdf.drawString(x_left, y, f"FECHA: {factura.fecha_factura.strftime('%d/%m/%Y')}")

    y_right = 710

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(x_right, y_right, "FACTURA GUBERNAMENTAL")

    pdf.setFont("Helvetica-Bold", 7)
    y_right -= 14
    pdf.drawString(x_right, y_right, f"NCF_{comprobante.ncf if comprobante else ''}")

    y_right -= 12
    pdf.drawString(x_right, y_right, f"VALIDO HASTA: {comprobante.fecha_validez.strftime('%d/%m/%Y') if comprobante else ''}")

    y = 575

    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawString(x_left, y, "CLIENTE :")
    pdf.drawString(x_left, y - 13, "RNC")

    pdf.drawString(160, y, factura.cliente_nombre.upper())
    pdf.drawString(160, y - 13, factura.cliente_rnc)

    y = 505

    pdf.setFont("Helvetica", 7)
    pdf.drawString(x_left, y, "Periodo de factura")

    periodo = f"Del {fecha_larga_es(factura.fecha_inicio)} al {fecha_larga_es(factura.fecha_fin)}"

    pdf.drawString(150, y, periodo)
    pdf.line(145, y - 2, 390, y - 2)

    y -= 26

    pdf.drawString(x_left, y, "Cantidad de conduces")
    pdf.line(150, y - 2, 245, y - 2)
    pdf.drawCentredString(197, y, str(factura.cantidad_conduces))

    pdf.drawString(265, y, "del No.")
    pdf.line(315, y - 2, 390, y - 2)
    pdf.drawCentredString(352, y, str(factura.conduce_inicial or ""))

    pdf.drawString(405, y, "al")
    pdf.line(430, y - 2, 505, y - 2)
    pdf.drawCentredString(467, y, str(factura.conduce_final or ""))

    y -= 14
    pdf.drawString(x_left, y, "Bloques")
    pdf.line(95, y - 2, 145, y - 2)
    pdf.drawCentredString(120, y, str(factura.bloques))

    data = [["PRODUCTO", "CANTIDAD", "PRECIO SIN ITEBIS", "VALOR RD$"]]

    for detalle in detalles:
        data.append([
            detalle.producto.upper(),
            formato_cantidad(detalle.cantidad),
            formato_monto(detalle.precio_sin_itbis),
            formato_monto(detalle.valor),
        ])

    table_x = margin_left
    table_top_y = 440

    col_producto = 120
    col_cantidad = 120
    col_precio = 170
    col_valor = content_width - col_producto - col_cantidad - col_precio

    row_heights = [18] + [26] * (len(data) - 1)
    table_height = sum(row_heights)
    table_bottom_y = table_top_y - table_height

    tabla = Table(data, colWidths=[col_producto, col_cantidad, col_precio, col_valor], rowHeights=row_heights)

    tabla.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 6.8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 6.8),
        ("ALIGN", (0, 1), (1, -1), "CENTER"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("RIGHTPADDING", (2, 1), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    tabla.wrapOn(pdf, width, height)
    tabla.drawOn(pdf, table_x, table_bottom_y)

    valor_col_x = table_x + col_producto + col_cantidad + col_precio
    valor_col_width = col_valor

    total_label_x = table_x + col_producto
    total_box_x = valor_col_x
    total_y = table_bottom_y - 24
    box_height = 24

    totales = [
        ("SUB-TOTAL PRODUCTOS EXENTOS", factura.subtotal_exento),
        ("SUB-TOTAL PRODUCTOS GRAVADOS", factura.subtotal_gravado),
        ("SUBTOTAL", factura.subtotal),
        ("ITBIS", factura.itbis),
        ("TOTAL", factura.total),
    ]

    for label, value in totales:
        pdf.setFont("Helvetica-Bold", 7)
        pdf.drawString(total_label_x, total_y + 8, label)

        pdf.setFillColor(colors.lightgrey)
        pdf.rect(total_box_x, total_y, valor_col_width, box_height, fill=1, stroke=1)
        pdf.setFillColor(colors.black)

        pdf.setFont("Helvetica-Bold", 7)
        pdf.drawRightString(total_box_x + valor_col_width - 5, total_y + 8, formato_monto(value))

        total_y -= box_height

    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawCentredString(width / 2, 155, "FIRMA Y SELLO DE LA EMPRESA")

    if factura.es_electronica and qrcode is not None:
        url_qr = factura.url_qr or construir_url_qr_ecf(factura)

        if url_qr:
            qr_img = qrcode.make(url_qr)
            qr_buffer = BytesIO()
            qr_img.save(qr_buffer, format="PNG")
            qr_buffer.seek(0)

            qr_reader = ImageReader(qr_buffer)

            qr_x = 58
            qr_y = 55
            qr_size = 95

            pdf.drawImage(qr_reader, qr_x, qr_y, qr_size, qr_size)

            pdf.setFont("Helvetica", 6)
            pdf.drawString(qr_x, qr_y - 10, f"Código de Seguridad: {factura.codigo_seguridad or ''}")

            if factura.fecha_firma_digital:
                pdf.drawString(qr_x, qr_y - 20, f"Fecha de Firma Digital: {factura.fecha_firma_digital.strftime('%d-%m-%Y %H:%M:%S')}")

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="factura_{factura.id}.pdf"'
    return response


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def editar_comprobante(request, comprobante_id):
    empresa = obtener_empresa(request)
    comprobante = get_object_or_404(ComprobanteFiscal, id=comprobante_id, empresa=empresa)

    if request.method == "POST":
        if comprobante.usado:
            messages.error(request, "No se puede editar un comprobante ya utilizado.")
            return redirect("facturacion")

        comprobante.ncf = request.POST.get("ncf")
        comprobante.tipo = request.POST.get("tipo")
        comprobante.fecha_validez = request.POST.get("fecha_validez")
        comprobante.save()

        messages.success(request, "Comprobante actualizado correctamente.")
        return redirect("facturacion")

    return redirect("facturacion")


@login_required(login_url="login_usuario")
@modulo_requerido("modulo_facturacion")
def eliminar_comprobante(request, comprobante_id):
    empresa = obtener_empresa(request)
    comprobante = get_object_or_404(ComprobanteFiscal, id=comprobante_id, empresa=empresa)

    if comprobante.usado:
        messages.error(request, "No se puede eliminar un comprobante ya utilizado.")
    else:
        comprobante.delete()
        messages.success(request, "Comprobante eliminado correctamente.")

    return redirect("facturacion")


# =====================================================
# AUTENTICACIÓN / REGISTRO / VALIDACIÓN
# =====================================================

def enviar_codigo_correo(user, tipo="correo"):
    CodigoValidacion.objects.filter(user=user, tipo=tipo, usado=False).update(usado=True)

    codigo = CodigoValidacion.objects.create(user=user, tipo=tipo)

    asunto = "Código de validación - SaaS Panaderías"

    mensaje = f"""
Hola {user.first_name or user.email},

Tu código de validación es:

{codigo.codigo}

Este código vence en 15 minutos.

Si no solicitaste este código, puedes ignorar este mensaje.
"""

    send_mail(asunto, mensaje, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)

    return codigo


def registro(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip().lower()
        password = request.POST.get("password", "")
        confirmar_password = request.POST.get("confirmar_password", "")

        empresa_nombre = request.POST.get("empresa_nombre", "").strip()
        empresa_rnc = request.POST.get("empresa_rnc", "").strip()
        empresa_telefono = request.POST.get("empresa_telefono", "").strip()

        # Compatibilidad con formularios anteriores
        if not correo:
            correo = request.POST.get("email", "").strip().lower()
        if not password:
            password = request.POST.get("password1", "")
        if not confirmar_password:
            confirmar_password = request.POST.get("password2", "")
        if not nombre:
            nombre = request.POST.get("username", "").strip()
        if not empresa_nombre:
            empresa_nombre = "Mi empresa"

        if not nombre or not correo or not password:
            messages.error(request, "Debe completar los campos obligatorios.")
            return redirect("registro")

        if password != confirmar_password:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect("registro")

        if User.objects.filter(username=correo).exists() or User.objects.filter(email=correo).exists():
            messages.error(request, "Ya existe una cuenta registrada con ese correo.")
            return redirect("registro")

        plan_basico = Plan.objects.filter(nombre__iexact="Básico").first()

        if not plan_basico:
            plan_basico = Plan.objects.create(
                nombre="Básico",
                precio=1500,
                limite_conduces=500,
                limite_usuarios=3,
                almacenamiento_gb=1
            )

        empresa_saas = EmpresaSaaS.objects.create(
            nombre=empresa_nombre,
            rnc=empresa_rnc or "",
            correo=correo,
            activa=True
        )

        user = User.objects.create(
            username=correo,
            email=correo,
            first_name=nombre,
            last_name=apellido,
            password=make_password(password),
            is_active=False
        )

        # Empresa operativa usada por conduces, centros, menús y facturas
        Empresa.objects.create(
            usuario=user,
            nombre=empresa_nombre,
            rnc=empresa_rnc or "",
            telefono=empresa_telefono or "",
            correo=correo,
            numero_inicial_conduce="0001",
        )

        PerfilUsuario.objects.create(
            user=user,
            empresa=empresa_saas,
            rol="admin_empresa",
            correo_validado=False,
            activo=True
        )

        Suscripcion.objects.create(
            empresa=empresa_saas,
            plan=plan_basico,
            estado="prueba",
            fecha_inicio=timezone.now().date(),
            fecha_fin=timezone.now().date() + timedelta(days=15),
            en_prueba=True
        )

        enviar_codigo_correo(user, tipo="correo")

        request.session["usuario_pendiente_id"] = user.id

        messages.success(request, "Cuenta creada. Te enviamos un código de validación al correo.")
        return redirect("verificar_correo")

    return render(request, "registro.html")


def verificar_correo(request):
    user_id = request.session.get("usuario_pendiente_id")

    if not user_id:
        messages.error(request, "No hay usuario pendiente de validación.")
        return redirect("login_usuario")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        codigo_ingresado = request.POST.get("codigo", "").strip()

        codigo = (
            CodigoValidacion.objects
            .filter(user=user, tipo="correo", codigo=codigo_ingresado, usado=False)
            .order_by("-creado_en")
            .first()
        )

        if not codigo or not codigo.esta_vigente():
            messages.error(request, "Código inválido o vencido.")
            return redirect("verificar_correo")

        codigo.usado = True
        codigo.save()

        user.is_active = True
        user.save()

        perfil = PerfilUsuario.objects.filter(user=user).first()
        if perfil:
            perfil.correo_validado = True
            perfil.save()
            request.session["empresa_id"] = perfil.empresa.id if perfil.empresa else None

        login(request, user)
        request.session.pop("usuario_pendiente_id", None)

        messages.success(request, "Correo validado correctamente. Bienvenido.")
        return redirect("inicio")

    return render(request, "verificar_correo.html", {"correo": user.email})


def reenviar_codigo_correo(request):
    user_id = request.session.get("usuario_pendiente_id")

    if not user_id:
        messages.error(request, "No hay usuario pendiente de validación.")
        return redirect("login_usuario")

    user = get_object_or_404(User, id=user_id)
    enviar_codigo_correo(user, tipo="correo")

    messages.success(request, "Te enviamos un nuevo código de validación.")
    return redirect("verificar_correo")


def login_usuario(request):
    if request.method == "POST":
        correo = (
            request.POST.get("correo", "") or
            request.POST.get("username", "") or
            request.POST.get("email", "")
        ).strip().lower()

        password = request.POST.get("password", "")
        remember = request.POST.get("remember_me")

        user = authenticate(request, username=correo, password=password)

        if user is None:
            messages.error(request, "Correo o contraseña incorrectos.")
            return redirect("login_usuario")

        if not user.is_active:
            request.session["usuario_pendiente_id"] = user.id
            enviar_codigo_correo(user, tipo="correo")
            messages.warning(request, "Debes validar tu correo. Te enviamos un nuevo código.")
            return redirect("verificar_correo")

        perfil = PerfilUsuario.objects.filter(user=user).first()

        if not perfil or not perfil.activo:
            messages.error(request, "Usuario inactivo. Contacte al administrador.")
            return redirect("login_usuario")

        login(request, user)

        if not remember:
            request.session.set_expiry(0)

        request.session["empresa_id"] = perfil.empresa.id if perfil.empresa else None

        return redirect("inicio")

    return render(request, "login.html")


def logout_usuario(request):
    logout(request)
    return redirect("login_usuario")

@login_required(login_url="login_usuario")
def mi_empresa(request):
    empresa = obtener_empresa(request)

    if not empresa:
        messages.error(request, "No se encontró una empresa asociada a este usuario.")
        return redirect("inicio")

    perfil = PerfilUsuario.objects.filter(user=request.user).first()

    if request.method == "POST":
        empresa.nombre = request.POST.get("nombre", "").strip()
        empresa.rnc = request.POST.get("rnc", "").strip()
        empresa.direccion = request.POST.get("direccion", "").strip()
        empresa.telefono = request.POST.get("telefono", "").strip()
        empresa.ciudad = request.POST.get("ciudad", "").strip()
        empresa.correo = request.POST.get("correo", "").strip()
        empresa.numero_inicial_conduce = request.POST.get("numero_inicial_conduce", "0001").strip()

        if request.FILES.get("logo"):
            empresa.logo = request.FILES.get("logo")

        empresa.modulo_conduces = request.POST.get("modulo_conduces") == "on"
        empresa.modulo_centros = request.POST.get("modulo_centros") == "on"
        empresa.modulo_menu = request.POST.get("modulo_menu") == "on"
        empresa.modulo_facturacion = request.POST.get("modulo_facturacion") == "on"
        empresa.modulo_reportes = request.POST.get("modulo_reportes") == "on"
        empresa.modulo_rutas = request.POST.get("modulo_rutas") == "on"
        empresa.modulo_nomina = request.POST.get("modulo_nomina") == "on"
        empresa.modulo_inventario = request.POST.get("modulo_inventario") == "on"

        empresa.save()

        messages.success(request, "Datos de la empresa actualizados correctamente.")
        return redirect("mi_empresa")

    usuarios = PerfilUsuario.objects.filter(
        empresa=perfil.empresa
    ).select_related("user") if perfil and perfil.empresa else []

    return render(request, "mi_empresa.html", {
        "empresa": empresa,
        "usuarios": usuarios,
        "perfil": perfil,
    })


@login_required(login_url="login_usuario")
def crear_usuario_empresa(request):
    empresa = obtener_empresa(request)
    perfil_actual = PerfilUsuario.objects.filter(user=request.user).first()

    if not perfil_actual or perfil_actual.rol != "admin_empresa":
        messages.error(request, "No tienes permisos para crear usuarios.")
        return redirect("mi_empresa")

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip().lower()
        password = request.POST.get("password", "").strip()
        rol = request.POST.get("rol", "consulta").strip()

        if not nombre or not correo or not password:
            messages.error(request, "Debe completar nombre, correo y contraseña.")
            return redirect("mi_empresa")

        if User.objects.filter(username=correo).exists() or User.objects.filter(email=correo).exists():
            messages.error(request, "Ya existe un usuario con ese correo.")
            return redirect("mi_empresa")

        user = User.objects.create(
            username=correo,
            email=correo,
            first_name=nombre,
            last_name=apellido,
            password=make_password(password),
            is_active=True
        )

        PerfilUsuario.objects.create(
            user=user,
            empresa=perfil_actual.empresa,
            rol=rol,
            correo_validado=True,
            activo=True
        )

        messages.success(request, "Usuario creado correctamente.")
        return redirect("mi_empresa")

    return redirect("mi_empresa")
@login_required(login_url="login_usuario")
def cartas_administrativas(request):
    empresa = obtener_empresa(request)

    if not empresa:
        messages.error(request, "No se encontró una empresa asociada.")
        return redirect("inicio")

    return render(request, "cartas_administrativas.html", {
        "empresa": empresa
    })


@login_required(login_url="login_usuario")
def generar_carta_pdf(request):
    empresa = obtener_empresa(request)

    if request.method != "POST":
        return redirect("cartas_administrativas")

    destinatario = request.POST.get("destinatario", "").strip()
    institucion = request.POST.get("institucion", "").strip()
    asunto = request.POST.get("asunto", "").strip()
    contenido = request.POST.get("contenido", "").strip()
    firmante = request.POST.get("firmante", "").strip()
    cargo = request.POST.get("cargo", "").strip()
    ciudad = request.POST.get("ciudad", "").strip() or empresa.ciudad or ""
    fecha = timezone.localdate()

    if not destinatario or not asunto or not contenido:
        messages.error(request, "Debe completar destinatario, asunto y contenido.")
        return redirect("cartas_administrativas")

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = 735

    # Logo solo para cartas administrativas
    if empresa.logo:
        try:
            logo = ImageReader(empresa.logo.path)
            pdf.drawImage(logo, 50, 705, width=75, height=75, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(width / 2, y, (empresa.nombre or "").upper())

    y -= 14
    pdf.setFont("Helvetica", 8)
    if empresa.direccion:
        pdf.drawCentredString(width / 2, y, empresa.direccion)

    y -= 11
    if ciudad:
        pdf.drawCentredString(width / 2, y, ciudad)

    y -= 11
    datos_contacto = []
    if empresa.correo:
        datos_contacto.append(f"Correo: {empresa.correo}")
    if empresa.telefono:
        datos_contacto.append(f"Teléfono: {empresa.telefono}")

    if datos_contacto:
        pdf.drawCentredString(width / 2, y, " | ".join(datos_contacto))

    y -= 11
    if empresa.rnc:
        pdf.drawCentredString(width / 2, y, f"RNC: {empresa.rnc}")

    y -= 45

    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(width - 55, y, f"{ciudad}, {fecha_larga_es(fecha)}")

    y -= 45

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(55, y, "Señores:")
    y -= 14

    pdf.setFont("Helvetica", 9)
    pdf.drawString(55, y, destinatario)
    y -= 14

    if institucion:
        pdf.drawString(55, y, institucion)
        y -= 14

    y -= 14

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(55, y, f"Asunto: {asunto}")
    y -= 30

    pdf.setFont("Helvetica", 9)

    estilo = ParagraphStyle(
        name="CartaContenido",
        fontName="Helvetica",
        fontSize=9,
        leading=14,
        alignment=TA_LEFT,
    )

    contenido_html = contenido.replace("\n", "<br/>")
    parrafo = Paragraph(contenido_html, estilo)

    ancho_texto = width - 110
    alto_disponible = y - 150
    _, alto_parrafo = parrafo.wrap(ancho_texto, alto_disponible)

    parrafo.drawOn(pdf, 55, y - alto_parrafo)

    y = y - alto_parrafo - 55

    pdf.setFont("Helvetica", 9)
    pdf.drawString(55, y, "Atentamente,")

    y -= 55

    pdf.line(55, y, 250, y)

    y -= 13
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(55, y, firmante or empresa.nombre or "")

    y -= 12
    pdf.setFont("Helvetica", 9)
    pdf.drawString(55, y, cargo or "Representante autorizado")

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="carta_administrativa.pdf"'
    return response